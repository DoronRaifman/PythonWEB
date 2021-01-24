from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active     # grab the active worksheet
ws1.title = "MultiplyTable"
for i in range(1, 10+1):
    for j in range(1, 10+1):
        colLetter = get_column_letter(j + 1)
        cell_address = f'{colLetter}{i + 1}'
        ws1[cell_address].value = i * j
wb.save("MultiplyTable.xlsx")

